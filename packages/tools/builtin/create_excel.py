"""create_excel — generate an Excel workbook from structured sheet data."""

from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING, Any

from pydantic import BaseModel, Field

from citnega.packages.protocol.callables.base import BaseCallable
from citnega.packages.protocol.callables.types import CallableType
from citnega.packages.tools.builtin._tool_base import ToolOutput, tool_policy

if TYPE_CHECKING:
    from citnega.packages.protocol.callables.context import CallContext


class ExcelSheet(BaseModel):
    name: str = Field(description="Sheet tab name.")
    headers: list[str] = Field(description="Column headers for the first row.")
    rows: list[list[Any]] = Field(
        default_factory=list,
        description="Data rows — each inner list maps positionally to headers.",
    )
    freeze_header: bool = Field(default=True, description="Freeze the header row.")
    auto_filter: bool = Field(default=True, description="Enable auto-filter on headers.")


class CreateExcelInput(BaseModel):
    filename: str = Field(description="Output file path, e.g. ~/Desktop/report.xlsx")
    sheets: list[ExcelSheet] = Field(description="One or more sheets to include in the workbook.")


class CreateExcelTool(BaseCallable):
    """Generate an Excel (.xlsx) workbook from one or more structured sheets."""

    name = "create_excel"
    description = (
        "Create an Excel (.xlsx) workbook from structured sheet data. "
        "Each sheet has a name, column headers, and data rows. "
        "Supports freeze pane and auto-filter. Returns the path of the created file."
    )
    callable_type = CallableType.TOOL
    input_schema = CreateExcelInput
    output_schema = ToolOutput
    policy = tool_policy(
        timeout_seconds=30.0,
        requires_approval=False,
        network_allowed=False,
    )

    async def _execute(self, input: CreateExcelInput, context: CallContext) -> ToolOutput:
        try:
            import openpyxl  # type: ignore[import-untyped]
            from openpyxl.styles import Font, PatternFill  # type: ignore[import-untyped]
        except ImportError:
            return ToolOutput(result="[create_excel: openpyxl not installed — run: pip install openpyxl]")

        out_path = Path(input.filename).expanduser().resolve()
        out_path.parent.mkdir(parents=True, exist_ok=True)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # remove default empty sheet

        header_font = Font(bold=True)
        header_fill = PatternFill(fill_type="solid", fgColor="4472C4")

        for sheet_data in input.sheets:
            ws = wb.create_sheet(title=sheet_data.name[:31])  # Excel tab name limit

            # Write headers
            for col_idx, header in enumerate(sheet_data.headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill

            # Write rows
            for row_idx, row in enumerate(sheet_data.rows, start=2):
                for col_idx, value in enumerate(row, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

            # Auto-fit column widths (approximate)
            for col in ws.columns:
                max_len = max((len(str(cell.value or "")) for cell in col), default=8)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

            if sheet_data.freeze_header:
                ws.freeze_panes = "A2"

            if sheet_data.auto_filter and sheet_data.headers:
                last_col = openpyxl.utils.get_column_letter(len(sheet_data.headers))
                ws.auto_filter.ref = f"A1:{last_col}1"

        try:
            wb.save(str(out_path))
        except Exception as exc:
            return ToolOutput(result=f"[create_excel: failed to write: {exc}]")

        total_rows = sum(len(s.rows) for s in input.sheets)
        return ToolOutput(
            result=(
                f"Excel created: {out_path}  "
                f"({len(input.sheets)} sheet(s), {total_rows} data row(s), {out_path.stat().st_size} bytes)"
            )
        )
